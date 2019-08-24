from openpyxl import Workbook
from openpyxl import load_workbook

#load mark.xlsx
wb = load_workbook('mark.xlsx')
#select mark.xlsx
sheet = wb.active

# d2 = sheet['D2'] to access D2 cell

#getting max rows
max_row = sheet.max_row
#getting max column
max_column = sheet.max_column
print(max_row)
print(max_column)

students = []

#inputing value into students array
#iterate over all cells
for i in range(2,max_row+1):
    # for j in range (0,max_column)
        student = [0,0,0,0,0,0,0,0,0,0]
        student[0]=sheet.cell(row=i,column=1).value
        student[1]=(sheet.cell(row=1,column=2).value,4,sheet.cell(row =i,column=2).value)
        student[2]=sheet.cell(row=1,column=3).value,4,sheet.cell(row =i,column=3).value
        student[3]=sheet.cell(row=1,column=4).value,3,sheet.cell(row =i,column=4).value
        student[4]=sheet.cell(row=1,column=5).value,1,sheet.cell(row =i,column=5).value
        student[5]=sheet.cell(row=1,column=6).value,1,sheet.cell(row =i,column=6).value
        student[6]=sheet.cell(row=1,column=7).value,4,sheet.cell(row =i,column=7).value
        student[7]=sheet.cell(row=1,column=8).value,3,sheet.cell(row =i,column=8).value
        student[8]=sheet.cell(row=1,column=9).value,3,sheet.cell(row =i,column=9).value
        student[9]=sheet.cell(row=1,column=10).value,1,sheet.cell(row =i,column=10).value
        students.append(student)

for stud in students:
    print(stud)

fail_count = 0
marks=[]
#traversing through each student
for student_item in students:
    total_mark = 0
    sgpa_total=0
    #traversing through each subject column of student_item
    for j in range (1,len(student_item)):
        if(student_item[j][2]=='F' or student_item[j][2] == 'FE'):
            fail_count+=1
        if(student_item[j][2]=='O'):
            total_mark+=10
            sgpa_total += 10 * student_item[j][1]
        elif(student_item[j][2]=='A+'):
            total_mark+=9
            sgpa_total += 9 * student_item[j][1]
        elif(student_item[j][2]=='A'):
            total_mark+=8.5
            sgpa_total += 8.5 * student_item[j][1]
        elif(student_item[j][2]=='B+'):
            total_mark+=8
            sgpa_total += 8 * student_item[j][1]
        elif(student_item[j][2]=='B'):
            total_mark+=7
            sgpa_total += 7 * student_item[j][1]
        elif(student_item[j][2]=='C'):
            total_mark+=6
            sgpa_total += 6 * student_item[j][1]
        elif(student_item[j][2]=='P'):
            total_mark+=5
            sgpa_total += 5 * student_item[j][1]
        else: pass
        
    marks.append((student_item[0],total_mark , sgpa_total/24))

print("fail count : ",fail_count)
for mark in marks :
    print("Register No. : ",mark[0],"\nTotal Mark : ",mark[1],"\nSGPA : ",mark[2],"\n\n")
